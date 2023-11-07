from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
import requests
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
spreadsheet_name = "attendance.xlsx"
slack_token = os.getenv("SLACK_TOKEN")

checkin_timestamps = {}
checkout_timestamps = {}
break_start_timestamps = {}
break_end_timestamps = {}

@app.route("/slack/command", methods=["POST"])
def slack_command():
    if request.form.get("token") == "n1jsHwt68Qo8mpd0p6SCoiqU":
        user_id = request.form.get("user_id")
        user_name, user_id = get_user_info(user_id)

        if user_name is None:
            return jsonify({"text": "User information not found."}), 404

        command = request.form.get("command")

        if command == "/checkin":
            if user_id in checkin_timestamps:
                return jsonify({"text": "You have already checked in. You can check out after 5 minutes."})

            timestamp = datetime.now()
            checkin_timestamps[user_id] = timestamp
            response_text = f"{user_name} ({user_id}) checked in at {timestamp}"
            update_attendance(user_name, "Check-In", timestamp)
        elif command == "/checkout":
            timestamp = datetime.now()
            if user_id not in checkin_timestamps:
                return jsonify({"text": "You must check in before checking out."})

            checkin_time = checkin_timestamps[user_id]
            # if timestamp - checkin_time < timedelta(minutes=5):
            #     return jsonify({"text": "You can only check out after 5 minutes of checking in."})

            del checkin_timestamps[user_id]
            checkout_timestamps[user_id] = timestamp

            response_text = f"{user_name} ({user_id}) checked out at {timestamp}."
            update_attendance(user_name, "Check-Out", timestamp)
        elif command == "/breakstart":
            if user_id not in checkin_timestamps:
                return jsonify({"text": "You must check in before starting your break."})

            if user_id in break_start_timestamps:
                return jsonify({"text": "You have already started your break. You can end it with '/breakend'."})

            timestamp = datetime.now()
            break_start_timestamps[user_id] = timestamp
            response_text = f"{user_name} ({user_id}) started a break at {timestamp}"
            update_attendance(user_name, "Break Start", timestamp)
        elif command == "/breakend":
            timestamp = datetime.now()
            if user_id not in checkin_timestamps:
                return jsonify({"text": "You must check in before ending your break."})

            if user_id not in break_start_timestamps:
                return jsonify({"text": "You must start your break with '/breakstart' before ending it."})

            del break_start_timestamps[user_id]
            response_text = f"{user_name} ({user_id}) ended the break at {timestamp}."
            update_attendance(user_name, "Break End", timestamp)
        else:
            response_text = "Invalid command."

        return jsonify({"text": response_text})
    else:
        return jsonify({"text": "Unauthorized request."}), 401

def get_user_info(user_id):
    headers = {"Authorization": f"Bearer {slack_token}"}
    response = requests.get(f"https://slack.com/api/users.info?user={user_id}", headers=headers)
    data = response.json()

    if response.status_code == 200 and data.get("ok", False):
        user_name = data["user"]["real_name"]
        return user_name, user_id
    else:
        print(f"Error fetching user information: {data}")
        return None, user_id

def update_attendance(user_name, action, timestamp):
    try:
        workbook = load_workbook(spreadsheet_name)
    except FileNotFoundError:
        workbook = Workbook()

    if user_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(title=user_name)
        worksheet.append(["Action", "Date", "Time"])
    else:
        worksheet = workbook[user_name]

    date = timestamp.date()
    time = timestamp.time()

    if action == "Check-In" or action == "Check-Out":
        worksheet.append([action, date, time])
    else:
        worksheet.append([action, date, time])

    workbook.save(spreadsheet_name)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=3000)
