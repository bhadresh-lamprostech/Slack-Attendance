from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
import requests
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
spreadsheet_name = "attendance.xlsx"
slack_token = os.getenv("SLACK_TOKEN")

checkin_timestamps = {}
checkout_timestamps = {}
break_start_timestamps = {}
break_end_timestamps = {}

slack_channel_id = "C063XRS7VKL"  # Replace with the ID of your Slack channel
bot_user_id = "U062M4APVGV"  # Replace with the ID of your Slack bot user

# Dictionary to store user IDs and their status (checked in or checked out)
user_status = {}
user_status_message_ts = None

@app.route("/slack/command", methods=["POST"])
def slack_command():
    if request.form.get("token") == "n1jsHwt68Qo8mpd0p6SCoiqU":
        user_id = request.form.get("user_id")
        user_name, user_id = get_user_info(user_id)

        if user_name is None:
            return jsonify({"text": "User information not found."}), 404

        command = request.form.get("command")

        if command == "/checkin":
            if user_id in checkin_timestamps:
                return jsonify({"text": "You have already checked in. You can check out after 5 minutes."})

            timestamp = datetime.now()
            checkin_timestamps[user_id] = timestamp
            response_text = f"{user_name} ({user_id}) checked in at {timestamp}"
            update_attendance(user_name, "Check-In", timestamp)

            user_status[user_id] = "online"
            update_slack_status(user_id, user_status[user_id])
        elif command == "/checkout":
            timestamp = datetime.now()
            if user_id not in checkin_timestamps:
                return jsonify({"text": "You must check in before checking out."})

            checkin_time = checkin_timestamps[user_id]

            del checkin_timestamps[user_id]
            checkout_timestamps[user_id] = timestamp

            response_text = f"{user_name} ({user_id}) checked out at {timestamp}."
            update_attendance(user_name, "Check-Out", timestamp)

            user_status[user_id] = "offline"
            update_slack_status(user_id, user_status[user_id])
        elif command == "/breakstart":
            if user_id not in checkin_timestamps:
                return jsonify({"text": "You must check in before starting your break."})

            if user_id in break_start_timestamps:
                return jsonify({"text": "You have already started your break. You can end it with '/breakend'."})
            
            user_status[user_id] = "offline"
            update_slack_status(user_id, user_status[user_id])
                

            timestamp = datetime.now()
            break_start_timestamps[user_id] = timestamp
            response_text = f"{user_name} ({user_id}) started a break at {timestamp}"
            update_attendance(user_name, "Break Start", timestamp)
        elif command == "/breakend":
            timestamp = datetime.now()
            if user_id not in checkin_timestamps:
                return jsonify({"text": "You must check in before ending your break."})

            if user_id not in break_start_timestamps:
                return jsonify({"text": "You must start your break with '/breakstart' before ending it."})
            user_status[user_id] = "online"
            update_slack_status(user_id, user_status[user_id])

            del break_start_timestamps[user_id]
            response_text = f"{user_name} ({user_id}) ended the break at {timestamp}."
            update_attendance(user_name, "Break End", timestamp)
        else:
            response_text = "Invalid command."

        user_status_list = create_user_status_list()
        update_status_message(user_status_list)

        return jsonify({"text": response_text})
    else:
        return jsonify({"text": "Unauthorized request."}), 401

def get_user_info(user_id):
    headers = {"Authorization": f"Bearer {slack_token}"}
    response = requests.get(f"https://slack.com/api/users.info?user={user_id}", headers=headers)
    data = response.json()

    if response.status_code == 200 and data.get("ok", False):
        user_name = data["user"]["real_name"]
        return user_name, user_id
    else:
        print(f"Error fetching user information: {data}")
        return None, user_id

def update_attendance(user_name, action, timestamp):
    try:
        workbook = load_workbook(spreadsheet_name)
    except FileNotFoundError:
        workbook = Workbook()

    if user_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(title=user_name)
        worksheet.append(["Action", "Date", "Time"])
    else:
        worksheet = workbook[user_name]

    date_str = timestamp.strftime("%Y-%m-%d")  # Format the date as "YYYY-MM-DD"
    time_str = timestamp.strftime("%H:%M:%S")  # Format the time as "HH:MM:SS"

    worksheet.append([action, date_str, time_str])

    workbook.save(spreadsheet_name)

def update_slack_status(user_id, status):
    headers = {"Authorization": f"Bearer {slack_token}"}
    data = {
        "token": slack_token,
        "user": user_id,
        "presence": status
    }
    response = requests.post("https://slack.com/api/users.setPresence", data=data)

def create_user_status_list():
    user_status_list = []

    for user_id, status in user_status.items():
        user_name, _ = get_user_info(user_id)
        indicator = ":large_green_circle:" if status == "online" else ":red_circle:"

        user_status_list.append(f"{user_name} {indicator}")

    return "\n".join(user_status_list)

def update_status_message(text):
    global user_status_message_ts

    if user_status_message_ts:
        headers = {
            "Authorization": f"Bearer {slack_token}",
            "Content-Type": "application/json"
        }
        message = {
            "channel": slack_channel_id,
            "text": text,
            "ts": user_status_message_ts
        }
        response = requests.post("https://slack.com/api/chat.update", headers=headers, json=message)

        if response.status_code == 200:
            return
        else:
            print(f"Error updating message in Slack. Status code: {response.status_code}")
            print(response.text)

    headers = {
        "Authorization": f"Bearer {slack_token}",
        "Content-Type": "application/json"
    }
    message = {
        "channel": slack_channel_id,
        "text": text,
        "as_user": bot_user_id
    }
    response = requests.post("https://slack.com/api/chat.postMessage", headers=headers, json=message)

    if response.status_code == 200:
        data = response.json()
        user_status_message_ts = data.get("ts")
    else:
        print(f"Error posting message to Slack. Status code: {response.status_code}")
        print(response.text)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=3000)
