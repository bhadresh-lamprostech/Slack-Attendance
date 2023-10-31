from openpyxl import load_workbook
from datetime import datetime, timedelta



def calculate_working_hours(sheet):
    data = {}
    for row in sheet.iter_rows(values_only=True):
        if row and len(row) >= 4:
            action, date, time, working_hours = row
            if action == 'Check-In':
                checkin = datetime.combine(date, time)
            elif action == 'Check-Out':
                checkout = datetime.combine(date, time)
                if checkin:
                    date_str = date.date().strftime('%Y-%m-%d')
                    if date_str not in data:
                        data[date_str] = {}
                    if sheet.title not in data[date_str]:
                        data[date_str][sheet.title] = working_hours
                    else:
                        data[date_str][sheet.title] += working_hours
                    checkin = None
    return data

def is_7_days_old(date):
    current_date = datetime.now().date()
    date_date = date.date()  # Extract the date part
    return (current_date - date_date).days >= 7

def delete_old_data_from_sheet(sheet):
    rows_to_delete = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= 2:  # Check for valid date and time columns
            date = row[1]  # Date is in the second column (column 'B')
            if date is not None:
                if is_7_days_old(date):
                    rows_to_delete.append(row)

    for row_data in rows_to_delete:
        # Find the row number and delete the entire row
        row_index = sheet.cell(row=2, column=1).row  # Assuming the header is in row 1
        sheet.delete_rows(row_index)

def main():
    spreadsheet_name = 'attendance.xlsx'
    workbook = load_workbook(spreadsheet_name)

    # Calculate working hours and organize data
    final_count_data = calculate_working_hours(workbook.active)

    # Delete old data from individual sheets
    for sheet_name in workbook.sheetnames:
        if sheet_name == 'FinalCount':
            continue
        sheet = workbook[sheet_name]
        delete_old_data_from_sheet(sheet)

    # Save the modified workbook
    workbook.save(spreadsheet_name)

if __name__ == '__main__':
    main()
